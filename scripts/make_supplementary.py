# -*- coding: utf-8 -*-
# Build the submission supplement: Supplementary_Tables.xlsx
# S1 full 731-trait MR results (conduction disorders) / S2 full 731 (AV block) / S3 the 15 cross-outcome concordant traits
# S4 sensitivity analyses (the 15 concordant traits) / S5 reverse MR / S6 instrument list / S7 leave-one-out for the sensitivity-stable traits
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Reads the local CSV copies and writes the supplement workbook
PROJ = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RES  = os.path.join(PROJ, "results", "tables")
PKG  = os.environ.get("SUBMISSION_DIR", os.path.join(PROJ, "supplementary"))
os.makedirs(PKG, exist_ok=True)

C  = pd.read_csv(os.path.join(RES, "MR_immune_CONDUCTIO.csv"))
A  = pd.read_csv(os.path.join(RES, "MR_immune_AVBLOCK.csv"))
S  = pd.read_csv(os.path.join(RES, "immune_sensitivity.csv"))
Rv = pd.read_csv(os.path.join(RES, "MR_reverse.csv"))
IV = pd.read_csv(os.path.join(RES, "immune_ivs.csv"))
LOO= pd.read_csv(os.path.join(RES, "immune_leaveoneout.csv"))
PS = pd.read_csv(os.path.join(RES, "immune_presso_steiger.csv"))
PW = pd.read_csv(os.path.join(RES, "power_analysis.csv"))
PWH= pd.read_csv(os.path.join(RES, "power_hits.csv"))
THR= pd.read_csv(os.path.join(RES, "threshold_sensitivity_summary.csv"))
GC = pd.read_csv(os.path.join(RES, "MR_immune_GWS_CONDUCTIO.csv"))
GA = pd.read_csv(os.path.join(RES, "MR_immune_GWS_AVBLOCK.csv"))
S18_PATH = os.path.join(RES, "S18_phenotype_power.csv")
S19_PATH = os.path.join(RES, "S19_coverage_summary.csv")
S20_PATH = os.path.join(RES, "S19_diagnostic_coverage.csv")
S21_PATH = os.path.join(RES, "S21_qc_summary.csv")
S22_PATH = os.path.join(RES, "S22_af_specificity.csv")
S23_PATH = os.path.join(RES, "S23_subendpoint_summary.csv")

# --- S1 / S2 full trait-level results, sorted by P, rounded to a sensible precision ---
# Some accessions show aggregate instrument behaviour that no reporting convention can produce
# (aggregate R2 > 1, or sum(F)/n > 1, which is invariant to rescaling of the reported effect
# sizes; Section 3.1). Their R2 and power metrics are not interpretable, so every table carrying a
# per-phenotype estimate flags them. The set is read from the S18 output rather than hard-coded
# here, so that changing the QC rule in 25_phenotype_power.R propagates to S1/S2 automatically.
if os.path.exists(S18_PATH):
    _pw = pd.read_csv(S18_PATH)
    _pw["QC flag"] = _pw["QC flag"].fillna("")
    QC_FLAGGED = dict(zip(_pw.loc[_pw["QC flag"] != "", "GWAS Catalog accession"],
                          _pw.loc[_pw["QC flag"] != "", "QC flag"]))
else:
    QC_FLAGGED = {}

def fmt_full(df):
    d = df.copy()
    if "p" in d:
        d = d.sort_values("p")  # sort on numeric P first, then format
    for c in ["b", "se"]:
        if c in d: d[c] = d[c].round(4)
    for c in ["OR", "OR_L", "OR_U"]:
        if c in d: d[c] = d[c].round(3)
    for c in ["p", "FDR"]:
        if c in d: d[c] = d[c].map(lambda x: f"{x:.2e}")
    d = d.rename(columns={
        "id": "GWAS ID", "trait": "Immune phenotype", "nIV": "N IVs",
        "method": "Method", "b": "beta", "se": "SE",
        "OR": "OR", "OR_L": "OR 95%CI lower", "OR_U": "OR 95%CI upper",
        "p": "P", "FDR": "FDR (BH)"})
    acc = d["GWAS ID"].str.replace("ebi-a-", "", regex=False)
    d["QC flag"] = acc.map(lambda a: f"{QC_FLAGGED[a]}; R2 and power metrics not interpretable"
                           if a in QC_FLAGGED else None)
    return d

S1 = fmt_full(C)
S2 = fmt_full(A)

# --- S3 the 15 direction-concordant traits across outcomes (as in Section 3.2) ---
m = C.merge(A[["id", "OR", "p"]], on="id", suffixes=("_C", "_A"))
conc = m[(m["p_C"] < 0.05) & (m["p_A"] < 0.05) &
         ((m["OR_C"] - 1) * (m["OR_A"] - 1) > 0)].copy().sort_values("p_C")
def lineage(t):
    t = t.lower()
    if "natural killer" in t or " nk" in t: return "NK cell"
    if "regulatory t cell" in t or "treg" in t: return "Regulatory T cell"
    if "dendritic" in t: return "Dendritic cell"
    if "granulocyte" in t: return "Granulocyte"
    if "monocyte" in t: return "Monocyte"
    if "b cell" in t: return "B cell"
    if "cd33" in t or "myeloid" in t: return "Myeloid"
    if "t cell" in t or "cd4" in t or "cd8" in t: return "T cell"
    return "Other"

S3 = pd.DataFrame({
    "Immune phenotype": conc["trait"], "Cell lineage": conc["trait"].map(lineage), "N IVs": conc["nIV"],
    "OR (conduction)": conc["OR_C"].round(3),
    "95% CI (conduction)": conc["OR_L"].round(3).astype(str) + "–" + conc["OR_U"].round(3).astype(str),
    "P (conduction)": conc["p_C"].map(lambda x: f"{x:.2e}"),
    "FDR (conduction)": conc["FDR"].map(lambda x: f"{x:.2e}"),
    "OR (AV block)": conc["OR_A"].round(3), "P (AV block)": conc["p_A"].map(lambda x: f"{x:.2e}")})

# --- S4 sensitivity analyses for the 15 concordant traits, with the robust flag and instrument strength (F, R2) ---
IV["F"] = (IV["beta"] / IV["se"]) ** 2
IV["R2"] = 2 * IV["eaf"] * (1 - IV["eaf"]) * (IV["beta"] ** 2)
if not {"mean_F", "min_F", "R2_pct"}.issubset(S.columns):
    fr2 = IV.groupby("trait").agg(mean_F=("F", "mean"), min_F=("F", "min"),
                                  R2_sum=("R2", "sum")).reset_index()
    S4 = S.merge(fr2, on="trait", how="left")
    S4["R2_pct"] = S4["R2_sum"] * 100
    S4 = S4.drop(columns=["R2_sum"])
else:
    S4 = S.copy()
S4["mean_F"] = S4["mean_F"].round(1)
S4["min_F"] = S4["min_F"].round(1)
S4["R2_pct"] = S4["R2_pct"].round(2)
if "I2" in S4.columns and "I2_GX" not in S4.columns:
    S4 = S4.rename(columns={"I2": "I2_GX"})
for c in ["IVW_OR", "WM_OR", "Egger_OR", "OR"]:
    if c in S4: S4[c] = S4[c].round(3)
for c in ["Egger_intercept"]:
    if c in S4: S4[c] = S4[c].round(4)
for c in ["IVW_p", "WM_p", "Egger_p", "Egger_intercept_p", "Q", "Q_p", "I2_GX"]:
    if c in S4: S4[c] = S4[c].round(4)
S4 = S4.rename(columns={"trait": "Immune phenotype", "nIV": "N IVs",
                        "mean_F": "Mean F", "min_F": "Min F", "R2_pct": "R2 (%)"})

# --- S5 reverse-direction MR ---
S5 = Rv.copy()
for c in ["OR", "OR_L", "OR_U"]:
    if c in S5: S5[c] = S5[c].round(3)
S5["p"] = S5["p"].round(3)
S5 = S5.rename(columns={"trait": "Immune phenotype", "nIV": "N IVs",
                        "method": "Method", "OR_L": "OR 95%CI lower",
                        "OR_U": "OR 95%CI upper", "p": "P", "note": "Note"})

# --- S6 instrument list, with the harmonised variant-outcome associations alongside ---
# STROBE-MR items 9a and 10a require both sides. Listing only the exposure side made those
# checklist entries inaccurate, so the harmonised outcome estimates written by
# 26_diagnostic_coverage.R are merged in here. They are aligned to the exposure effect allele.
S6 = IV.copy()
_harm_path = os.path.join(RES, "harmonised_outcome_associations.csv")
if os.path.exists(_harm_path):
    _h = pd.read_csv(_harm_path)
    S6 = S6.merge(_h, on=["rsid", "id"], how="left")
    for _c in [c for c in S6.columns if c.startswith(("beta_", "se_", "eaf_"))]:
        S6[_c] = S6[_c].round(5)
    for _c in [c for c in S6.columns if c.startswith("p_")]:
        S6[_c] = S6[_c].map(lambda x: f"{x:.3g}" if pd.notna(x) else "")
S6 = S6.rename(columns={
    "id": "GWAS ID", "trait": "Immune phenotype", "chr": "CHR",
    "position": "POS", "rsid": "rsID", "ea": "EA", "nea": "NEA",
    "eaf": "EAF", "beta": "beta", "se": "SE", "p": "P", "n": "N",
    "beta_CONDUCTIO": "beta (conduction, harmonised)", "se_CONDUCTIO": "SE (conduction)",
    "p_CONDUCTIO": "P (conduction)", "eaf_CONDUCTIO": "EAF (conduction, FinnGen)",
    "beta_AVBLOCK": "beta (AV block, harmonised)", "se_AVBLOCK": "SE (AV block)",
    "p_AVBLOCK": "P (AV block)", "eaf_AVBLOCK": "EAF (AV block, FinnGen)"})

# --- S7 leave-one-out for the sensitivity-stable traits ---
# S4 and S7 both cover the 15 cross-outcome concordant phenotypes; restricting S7 to the
# sensitivity-stable subset discarded rows that immune_leaveoneout.csv already contains.
S7 = LOO[LOO["trait"].isin(set(S["trait"]))].copy()
S7["OR"] = S7["OR"].round(3)
S7["p"] = S7["p"].map(lambda x: f"{x:.2e}")
S7 = S7.rename(columns={"trait": "Immune phenotype", "drop_rsid": "SNP removed", "p": "P"})

# --- S8 MR-PRESSO and Steiger for the 5 sensitivity-stable traits ---
S8 = PS.rename(columns={
    "trait": "Immune phenotype", "nIV": "N IVs", "meanF": "Mean F",
    "R2_exp_pct": "R2 exposure (%)", "R2_out_pct": "R2 outcome (%)",
    "steiger_prop_correct": "Steiger prop. correct", "steiger_p": "Steiger P",
    "steiger_dir": "Steiger direction", "presso_global_p": "MR-PRESSO global P",
    "presso_n_outlier": "MR-PRESSO outliers", "OR_raw": "OR (all IVs)",
    "OR_adj": "OR (outlier-corrected)"})
# MRPRESSO runs the variant-level outlier test only when the global test is significant, so
# both columns are empty for this subset; label them rather than leaving blank cells.
ns = S8["MR-PRESSO global P"] >= 0.05
for c in ("MR-PRESSO outliers", "OR (outlier-corrected)"):
    S8[c] = S8[c].astype(object)
S8.loc[ns & S8["MR-PRESSO outliers"].isna(),
       "MR-PRESSO outliers"] = "not performed (global test not significant)"
S8.loc[ns & S8["OR (outlier-corrected)"].isna(),
       "OR (outlier-corrected)"] = "not applicable"

# --- S9 power: minimum detectable OR at 80% power across values of R2 ---
S9 = PW.rename(columns={
    "outcome": "Outcome", "R2_pct": "R2 (%)",
    "minDetectableOR_nominal": "Min detectable OR (nominal α=0.05)",
    "minDetectableOR_studywide": "Min detectable OR (study-wide α=0.05/731)"})

# --- S10 realised power of the five sensitivity-stable phenotypes at their observed effects ---
S10 = PWH.rename(columns={
    "trait": "Immune phenotype", "R2_exp_pct": "R2 exposure (%)", "OR": "Observed OR",
    "power_nominal": "Power (nominal α=0.05)", "power_studywide": "Power (study-wide α=0.05/731)"})

# --- S11 genome-wide instrument-threshold sensitivity analysis ---
Gtop = GC.sort_values("FDR").head(20).merge(
    GA[["id", "OR", "p", "FDR"]], on="id", how="left", suffixes=("_CONDUCTIO", "_AVBLOCK"))
for c in ["OR_CONDUCTIO", "OR_L", "OR_U", "OR_AVBLOCK"]:
    if c in Gtop:
        Gtop[c] = Gtop[c].round(3)
for c in ["p_CONDUCTIO", "FDR_CONDUCTIO", "p_AVBLOCK", "FDR_AVBLOCK"]:
    if c in Gtop:
        Gtop[c] = Gtop[c].map(lambda x: "" if pd.isna(x) else f"{x:.2e}")
for c in ["mean_F", "min_F"]:
    if c in Gtop:
        Gtop[c] = Gtop[c].round(1)
Gtop = Gtop.rename(columns={
    "id": "GWAS ID", "trait": "Immune phenotype", "nIV": "N IVs",
    "method": "Method", "mean_F": "Mean F", "min_F": "Min F",
    "OR_CONDUCTIO": "OR (conduction)", "OR_L": "OR 95%CI lower (conduction)",
    "OR_U": "OR 95%CI upper (conduction)",
    "p_CONDUCTIO": "P (conduction)", "FDR_CONDUCTIO": "FDR (conduction)",
    "OR_AVBLOCK": "OR (AV block)", "p_AVBLOCK": "P (AV block)",
    "FDR_AVBLOCK": "FDR (AV block)"})
S11a = THR.rename(columns={
    "threshold": "Exposure-IV threshold",
    "exposure_iv_records": "Exposure IV records",
    "traits_with_CONDUCTIO_results": "Traits with conduction results",
    "CONDUCTIO_nominal": "Conduction P<0.05",
    "CONDUCTIO_FDR": "Conduction FDR<0.05",
    "CONDUCTIO_min_FDR": "Minimum conduction FDR",
    "traits_with_AVBLOCK_results": "Traits with AV block results",
    "AVBLOCK_nominal": "AV block P<0.05",
    "AVBLOCK_FDR": "AV block FDR<0.05",
    "AVBLOCK_min_FDR": "Minimum AV block FDR"})
S11a["Minimum conduction FDR"] = S11a["Minimum conduction FDR"].map(lambda x: f"{x:.3g}")
S11a["Minimum AV block FDR"] = S11a["Minimum AV block FDR"].map(lambda x: f"{x:.3g}")
S11 = pd.concat([
    S11a,
    pd.DataFrame([{}]),
    pd.DataFrame({"Exposure-IV threshold": ["Top 20 conduction signals under P<5e-8 threshold"]}),
    Gtop
], ignore_index=True)

# --- S12 trait-level cross-threshold comparison (Section 3.2): P and FDR under both thresholds ---
S12 = pd.read_csv(os.path.join(PROJ, "derived", "threshold_instability.csv"))
S12 = S12.sort_values(["outcome", "p_gws"]).rename(columns={
    "outcome": "Outcome",
    "id": "GWAS ID",
    "trait": "Immune phenotype",
    "nIV_main": "N IVs (P<1e-5)",
    "p_main": "P (P<1e-5)",
    "FDR_main": "FDR (P<1e-5)",
    "nIV_gws": "N IVs (P<5e-8)",
    "p_gws": "P (P<5e-8)",
    "FDR_gws": "FDR (P<5e-8)",
})
# Effect estimates under both thresholds are attached here. Without them the supplement reported
# only P and FDR for the genome-wide threshold, so 587 of the 607 phenotype-level effect sizes
# were absent from the submitted material even though they exist in the result tables.
_eff = []
for _oc, _main, _gws in [("Cardiac conduction disorders", C, GC),
                         ("Atrioventricular block", A, GA)]:
    _m = _main[["id", "OR", "OR_L", "OR_U"]].rename(
        columns={"OR": "OR (P<1e-5)", "OR_L": "_ml", "OR_U": "_mu"})
    _g = _gws[["id", "OR", "OR_L", "OR_U"]].rename(
        columns={"OR": "OR (P<5e-8)", "OR_L": "_gl", "OR_U": "_gu"})
    _d = _m.merge(_g, on="id", how="outer")
    _d["Outcome"] = _oc
    _eff.append(_d)
_eff = pd.concat(_eff, ignore_index=True).rename(columns={"id": "GWAS ID"})
S12 = S12.merge(_eff, on=["Outcome", "GWAS ID"], how="left")
def _ci(row, lo, hi):
    if pd.isna(row[lo]) or pd.isna(row[hi]):
        return ""
    return f"{row[lo]:.3f}-{row[hi]:.3f}"
S12["95% CI (P<1e-5)"] = S12.apply(lambda r: _ci(r, "_ml", "_mu"), axis=1)
S12["95% CI (P<5e-8)"] = S12.apply(lambda r: _ci(r, "_gl", "_gu"), axis=1)
S12 = S12.drop(columns=["_ml", "_mu", "_gl", "_gu"])
for _c in ["OR (P<1e-5)", "OR (P<5e-8)"]:
    S12[_c] = S12[_c].map(lambda x: round(x, 3) if pd.notna(x) else "")
for _c in ["P (P<1e-5)", "FDR (P<1e-5)", "P (P<5e-8)", "FDR (P<5e-8)"]:
    S12[_c] = S12[_c].map(lambda x: f"{x:.3g}" if pd.notna(x) else "")
for _c in ["N IVs (P<1e-5)", "N IVs (P<5e-8)"]:
    S12[_c] = S12[_c].map(lambda x: int(x) if pd.notna(x) else "")
S12 = S12[["Outcome", "GWAS ID", "Immune phenotype",
           "N IVs (P<1e-5)", "OR (P<1e-5)", "95% CI (P<1e-5)", "P (P<1e-5)", "FDR (P<1e-5)",
           "N IVs (P<5e-8)", "OR (P<5e-8)", "95% CI (P<5e-8)", "P (P<5e-8)", "FDR (P<5e-8)"]]

# --- S17 endpoint definition transparency table ---
S17 = pd.DataFrame([
    {
        "Endpoint": "Cardiac conduction disorders",
        "FinnGen code": "I9_CONDUCTIO",
        "Clinical construct": "Conduction-system disease spectrum",
        "Included diagnostic codes": "ICD-10 I44, I45; ICD-9 426 (Finnish health registers and cause-of-death register)",
        "Included sub-endpoints": "I9_AVBLOCK, I9_LBBB, I9_RBBB",
        "Total cases in GWAS": "12,371",
        "Controls in GWAS": "342,690",
        "Public exclusion rules": "Endpoint-level inclusion and exclusion rules are provided on the version-pinned Risteys page; mutually exclusive component-level exclusions are not reported in the public summary data.",
        "Component case counts": "Not reported at mutually exclusive component level in public summary data",
        "Pacemaker implantation proportion": "Not available in public summary data",
    },
    {
        "Endpoint": "Atrioventricular block",
        "FinnGen code": "I9_AVBLOCK",
        "Clinical construct": "Atrioventricular block, first to third degree",
        "Included diagnostic codes": "ICD-10 I44.0, I44.1, I44.2, I44.3; ICD-9 426[0-1] (Finnish health registers and cause-of-death register)",
        "Included sub-endpoints": "None",
        "Total cases in GWAS": "6,935",
        "Controls in GWAS": "342,690",
        "Public exclusion rules": "Endpoint-level inclusion and exclusion rules are provided on the version-pinned Risteys page; mutually exclusive component-level exclusions are not reported in the public summary data.",
        "Component case counts": "Not reported at mutually exclusive component level in public summary data",
        "Pacemaker implantation proportion": "Not available in public summary data",
    },
    {},
    {
        "Endpoint": "Endpoint definitions were taken from the version-pinned Risteys release 11 pages for I9_CONDUCTIO and I9_AVBLOCK (https://r11.risteys.finngen.fi/endpoints/I9_CONDUCTIO and https://r11.risteys.finngen.fi/endpoints/I9_AVBLOCK), accessed 16 July 2026.",
    },
    {
        "Endpoint": "Sinus-node disease enters the broader endpoint only as sinoatrial block and sinus arrest, which are coded within I45.5 (ICD-9 426.6) and therefore contribute cases. Sick sinus syndrome is coded as I49.5, lies outside the I44-I45 range, and does not contribute to either endpoint. These two should not be conflated.",
    },
    {
        "Endpoint": "Because no sub-code exclusions are applied, the broader endpoint also includes diagnoses whose mechanism is not degenerative conduction-system disease: pre-excitation syndrome (I45.6; ICD-9 426.7) and the Lown-Ganong-Levine syndrome (ICD-9 426.81), both arising from a congenital accessory pathway; the long QT syndrome (I45.8; ICD-9 426.82), a repolarization channelopathy rather than a conduction defect; and atrioventricular dissociation (ICD-9 426.89). The share of cases contributed by each sub-code is not resolvable from the public summary data.",
    },
    {
        "Endpoint": "The release 11 rules for I9_CONDUCTIO give HD_ICD_10 = I44|I45, COD_ICD_10 = I44|I45, HD_ICD_9 = 426 and INCLUDE = I9_AVBLOCK|I9_LBBB|I9_RBBB, with both HD_ICD_10_EXCL and COD_ICD_10_EXCL empty. Because no sub-code exclusions are applied, all ICD-10 I45 sub-codes are included, among them pre-excitation syndrome (I45.6), which has a congenital accessory-pathway mechanism distinct from degenerative conduction disease.",
    },
    {
        "Endpoint": "Total cases and controls are the counts remaining after genotype quality control, as reported at the corresponding step of those pages; these correspond to the GWAS summary statistics analysed here. The same pages report pre-quality-control counts at the registry-filter step of 12,786 for I9_CONDUCTIO and 7,173 for I9_AVBLOCK.",
    },
    {
        "Endpoint": "ICD = International Classification of Diseases, GWAS = genome-wide association study.",
    },
], columns=[
    "Endpoint", "FinnGen code", "Clinical construct", "Included diagnostic codes",
    "Included sub-endpoints", "Total cases in GWAS", "Controls in GWAS",
    "Public exclusion rules", "Component case counts", "Pacemaker implantation proportion"
])

# --- S18 phenotype-level instrument strength and power, by outcome ---
# Read the output of scripts/25_phenotype_power.R rather than recomputing here. That script
# reuses the harmonisation block of 12b_immune_mr.R, so its instrument counts match S1/S2
# exactly; recomputing from the unharmonised immune_ivs.csv does not (it disagreed for 531
# of 731 phenotypes) and must not be reintroduced.
if os.path.exists(S18_PATH):
    S18 = pd.read_csv(S18_PATH)
    # keep numeric cells numeric so the workbook formats them as numbers, while leaving the
    # "not interpretable" placeholders as text
    for col in S18.columns:
        if col in ("Outcome", "Immunophenotype", "GWAS Catalog accession", "QC flag"):
            continue
        S18[col] = S18[col].map(lambda v: pd.to_numeric(v, errors="coerce")
                                if pd.notna(pd.to_numeric(v, errors="coerce")) else v)
    S18 = pd.concat([S18, pd.DataFrame([{}]), pd.DataFrame([
        {"Outcome": "Instrument counts and R2 were computed on the harmonised instrument set used by the primary MR analysis (scripts 12b and 25), so they match the counts reported in Supplementary Tables S1 and S2 exactly."},
        {"Outcome": "Power and minimum detectable odds ratios were computed with the binary-outcome approximation of Brion et al, se = 1/sqrt(N x R2 x K x (1 - K)), evaluated at the study-wide threshold alpha = .05/731."},
        {"Outcome": "R2 values were estimated in the same discovery GWAS used for instrument selection and are therefore expected to be optimistic because of winner's curse; power values should be interpreted as approximate upper-bound estimates."},
        {"Outcome": "The QC flag marks phenotypes whose aggregate instrument behaviour is implausible on at least one of two criteria: aggregate R2 above 1, which is impossible for a variance-standardised exposure, and sum(F)/n above 1, which is the scale-invariant counterpart of that quantity and cannot be produced by any rescaling of the reported effect sizes. Power is not computed for flagged phenotypes and they are excluded from the summary statistics reported in Section 3.1."},
        {"Outcome": "OR = odds ratio, QC = quality control."},
    ])], ignore_index=True)
else:
    S18 = pd.DataFrame({"Outcome": ["S18_phenotype_power.csv was not found; run scripts/25_phenotype_power.R <finngen_dir>."]})

# --- S13 targeted re-analysis of the reported lymphocyte-count association ---
S13_PATH = os.path.join(RES, "S13_chen_reanalysis.csv")
if os.path.exists(S13_PATH):
    _c = pd.read_csv(S13_PATH)
    _c["outcome"] = _c["outcome"].map({"AVBLOCK": "Atrioventricular block",
                                       "CONDUCTIO": "Cardiac conduction disorders"})
    _c = _c.sort_values("outcome")
    for _col in ["minF", "medF"]:
        _c[_col] = _c[_col].round(0).astype(int)
    for _col in ["OR", "L", "U", "WM_OR", "EG_OR"]:
        _c[_col] = _c[_col].round(3)
    for _col in ["P", "WM_P", "EG_P", "EG_int", "EG_int_P", "Q", "Q_P"]:
        _c[_col] = _c[_col].round(5)
    S13 = _c.rename(columns={
        "outcome": "Outcome (FinnGen R11)", "nIV": "N instruments",
        "minF": "Min F", "medF": "Median F", "OR": "IVW OR",
        "L": "IVW 95%CI lower", "U": "IVW 95%CI upper", "P": "IVW P",
        "WM_OR": "Weighted-median OR", "WM_P": "Weighted-median P",
        "EG_OR": "MR-Egger OR", "EG_P": "MR-Egger P",
        "EG_int": "MR-Egger intercept", "EG_int_P": "MR-Egger intercept P",
        "Q": "Cochran's Q", "Q_df": "Q df", "Q_P": "Q P"})
else:
    S13 = pd.DataFrame({"Outcome (FinnGen R11)": ["S13_chen_reanalysis.csv was not found; run scripts/22_chen_reanalysis.R."]})

# --- S14 / S15 / S16 post-hoc rare-variant diagnostics, built from the CSVs written by
# 23_rare_variant_diagnostics.py. These were previously carried over from a pre-existing
# workbook, which meant a clean checkout could not rebuild them; they are read from disk
# here so that the workbook is reproducible from the result tables alone.
POSTHOC = {
    "S14": ("S14_rare_variant_diagnostics.csv", "S14 Rare-variant diagnostics",
            "Supplementary Table S14. Variant-level composition and post-hoc diagnostics for the five phenotypes reaching FDR significance under the genome-wide instrument threshold (P < 5 x 10-8); outcome, cardiac conduction disorders (FinnGen R11 I9_CONDUCTIO). Weights are shares of the inverse-variance weight. Post-hoc analysis; not prespecified."),
    "S15": ("S15_maf_sensitivity.csv", "S15 MAF sensitivity",
            "Supplementary Table S15. Post-hoc minor-allele-frequency sensitivity analysis. Both instrument thresholds were repeated after excluding instruments with MAF < 1%; records with no reported effect-allele frequency were also excluded because their MAF could not be confirmed. Post-hoc analysis; not prespecified."),
    "S16": ("S16_cutpoint_family.csv", "S16 Cut-point family",
            "Supplementary Table S16. Post-hoc exclusion of low-frequency instruments repeated across a family of cut-offs. MAF is min(EAF, 1 - EAF) from the effect-allele frequency reported in the exposure dataset; MAC is 2 x n x MAF using the analysis sample size reported for each trait. Records with no reported effect-allele frequency are excluded throughout because their MAF cannot be confirmed. Post-hoc analysis; not prespecified."),
}
POSTHOC_SHEETS = []
for _key, (_csv, _title, _cap) in POSTHOC.items():
    _path = os.path.join(RES, _csv)
    if os.path.exists(_path):
        _df = pd.read_csv(_path)
        _df.columns = [c.replace("_", " ") for c in _df.columns]
    else:
        _df = pd.DataFrame({"Note": [f"{_csv} was not found; run scripts/23_rare_variant_diagnostics.py <finngen_dir>."]})
    POSTHOC_SHEETS.append((_title, _df, _cap))

# --- S19 / S20 which diagnostics are computable at each threshold, and per-phenotype Q ---
# A diagnostic that cannot be computed is not the same as a diagnostic that was passed, so the
# instrument-count requirements are reported explicitly rather than left implicit.
if os.path.exists(S19_PATH):
    S19 = pd.read_csv(S19_PATH)
    S19 = pd.concat([S19, pd.DataFrame([{}]), pd.DataFrame([
        {"Outcome": "Minimum instrument counts: Cochran's Q and MR-Egger require at least three variants, MR-PRESSO at least four, and leave-one-out is uninformative below three. A single-variant Wald ratio supports none of them."},
        {"Outcome": "Cochran's Q is reported under two weightings. The first-order weight assumes the variant-exposure coefficients are measured without error. The modified second-order weight of Bowden et al (Int J Epidemiol 2019) propagates the variance of those coefficients and is evaluated iteratively at the inverse-variance-weighted estimate. In this dataset the two give nearly identical results, so the heterogeneity reported here is not an artefact of the first-order assumption."},
        {"Outcome": "Cochran's Q was computed for every phenotype meeting the three-variant requirement; per-phenotype values are given in Supplementary Table S20."},
        {"Outcome": "IV = instrumental variable."},
    ])], ignore_index=True)
else:
    S19 = pd.DataFrame({"Outcome": ["S19_coverage_summary.csv was not found; run scripts/26_diagnostic_coverage.R <finngen_dir>."]})

if os.path.exists(S20_PATH):
    S20 = pd.read_csv(S20_PATH)
    for c in ("Cochran Q", "Q P"):
        if c in S20:
            S20[c] = S20[c].round(4)
else:
    S20 = pd.DataFrame({"Outcome": ["S19_diagnostic_coverage.csv was not found; run scripts/26_diagnostic_coverage.R <finngen_dir>."]})

if os.path.exists(S21_PATH):
    S21 = pd.read_csv(S21_PATH)
    S21 = pd.concat([S21, pd.DataFrame([{}]), pd.DataFrame([
        {"Outcome": "Instruments were removed if the reported effect exceeded one standard deviation per allele, which is not credible for an inverse-normal transformed exposure, or if the expected minor-allele count (2 x n x MAF) was below 20. The filter retained 15,251 of 18,728 instrument records (81.4%). The two criteria overlap heavily: almost all of the removed records fail the effect-size criterion."},
        {"Outcome": "Records with no reported effect-allele frequency have no defined minor-allele count and were also removed."},
        {"Outcome": "The filter acts on individual variants rather than on phenotype-level aggregates, and is therefore complementary to the aggregate criteria used in Supplementary Table S18."},
        {"Outcome": "Per-phenotype results are in S21_instrument_qc_sensitivity.csv in the code package. Post-hoc analysis; not prespecified."},
    ])], ignore_index=True)
else:
    S21 = pd.DataFrame({"Outcome": ["S21_qc_summary.csv was not found; run scripts/27_instrument_qc_sensitivity.R <finngen_dir>."]})

# --- S22 / S23 specificity control and component sub-endpoints ---
if os.path.exists(S22_PATH):
    S22 = pd.read_csv(S22_PATH)
    S22 = pd.concat([S22, pd.DataFrame([{}]), pd.DataFrame([
        {"Outcome": "The same 232 lymphocyte-count instruments were applied to all three endpoints in FinnGen release 11, with identical harmonisation. Atrial fibrillation is included because its treatment can generate conduction-disorder codes, so an exposure acting on atrial fibrillation could appear to act on conduction disease."},
        {"Outcome": "Post-hoc analysis; not prespecified."},
    ])], ignore_index=True)
else:
    S22 = pd.DataFrame({"Outcome": ["S22_af_specificity.csv was not found; run scripts/28_af_specificity.R <chen_outdir> <finngen_dir>."]})

if os.path.exists(S23_PATH):
    S23 = pd.read_csv(S23_PATH)
    S23 = pd.concat([S23, pd.DataFrame([{}]), pd.DataFrame([
        {"Outcome": "The primary screen was repeated without modification on the two bundle-branch-block component endpoints of I9_CONDUCTIO. These are narrower than the broader endpoint and do not carry the pre-excitation, long-QT or sinoatrial-block sub-codes discussed in Section 2.3."},
        {"Outcome": "Case counts are 2,419 for I9_LBBB and 1,186 for I9_RBBB, against 12,371 and 6,935 for the primary endpoints, with the same 342,690 controls. The downloaded FinnGen release 11 summary files carry no case-count field, so these were recovered from the reported overall, case-only and control-only alternate-allele frequencies (scripts/30_endpoint_case_counts.R); the same procedure reproduced the two known case counts exactly."},
        {"Outcome": "Per-phenotype results are in S23_subendpoint_screen.csv in the code package. Post-hoc analysis; not prespecified."},
    ])], ignore_index=True)
else:
    S23 = pd.DataFrame({"Outcome": ["S23_subendpoint_summary.csv was not found; run scripts/29_subendpoint_screen.R <finngen_dir>."]})

SHEETS = [
    ("S1 Conduction full", S1,
     "Supplementary Table S1. Full two-sample MR results for all 731 immune cell phenotypes on cardiac conduction disorders (FinnGen R11 I9_CONDUCTIO; 12,371 cases/342,690 controls). IVW primary; sorted by P. Phenotypes whose aggregate instrument behaviour is implausible (aggregate R2 above 1, or the scale-invariant quantity sum(F)/n above 1) are flagged in the QC column. Their R2 and power metrics are not interpretable, and those with aggregate R2 above 1 are excluded from the power summaries; the MR estimates themselves are retained in the primary screen (Section 3.1)."),
    ("S2 AVblock full", S2,
     "Supplementary Table S2. Full two-sample MR results for all 731 immune cell phenotypes on atrioventricular block (FinnGen R11 I9_AVBLOCK; 6,935 cases/342,690 controls). IVW primary; sorted by P. Phenotypes whose aggregate instrument behaviour is implausible (aggregate R2 above 1, or the scale-invariant quantity sum(F)/n above 1) are flagged in the QC column. Their R2 and power metrics are not interpretable, and those with aggregate R2 above 1 are excluded from the power summaries; the MR estimates themselves are retained in the primary screen (Section 3.1)."),
    ("S3 Concordant", S3,
     "Supplementary Table S3. Fifteen immunophenotypes nominally significant (P<0.05) and directionally concordant across both outcomes (source of Section 3.2)."),
    ("S4 Sensitivity", S4,
     "Supplementary Table S4. Sensitivity analyses (IVW, weighted median, MR-Egger, MR-Egger I2_GX/NOME, and Cochran's Q), instrument strength (mean/min F), and variance explained (R2, %) for the 15 concordant immunophenotypes. robust=TRUE marks the five sensitivity-stable phenotypes (Section 3.4)."),
    ("S5 Reverse MR", S5,
     "Supplementary Table S5. Reverse Mendelian randomization (cardiac conduction disorders to immune phenotype) for the five sensitivity-stable phenotypes."),
    ("S6 Instruments", S6,
     "Supplementary Table S6. Genetic instruments (SNPs) used for all immune cell phenotype exposures (P<1e-5, r2<0.001, F>10), with the harmonised variant-outcome associations for both endpoints alongside the variant-exposure associations. Outcome estimates are aligned to the exposure effect allele; they are blank for the 1,053 instrument records that were dropped at harmonisation (allele mismatch, ambiguous palindromic variant, or absence from the outcome file), which is why 17,675 of the 18,728 records carry outcome statistics. GWAS ID should be used with phenotype label as the unique exposure key because two phenotype labels recur across GWAS accessions."),
    ("S7 Leave-one-out", S7,
     "Supplementary Table S7. Leave-one-out IVW estimates for the 15 cross-outcome concordant immunophenotypes for the primary outcome."),
    ("S8 PRESSO-Steiger", S8,
     "Supplementary Table S8. MR-PRESSO global test, outlier fields reported for transparency only, and Steiger directionality test (exposure vs outcome variance explained) for the five sensitivity-stable phenotypes and primary outcome."),
    ("S9 Power grid", S9,
     "Supplementary Table S9. Statistical power: minimum detectable odds ratio at 80% power as a function of instrument-explained variance (R²), for each outcome, at nominal (α=0.05) and study-wide (α=0.05/731) significance (Brion 2013 approximation)."),
    ("S10 Power per-hit", S10,
     "Supplementary Table S10. Statistical power to detect the observed effect of each of the five sensitivity-stable phenotypes (Section 3.4; primary outcome), given its instrument R² and observed OR, at nominal and study-wide significance."),
    ("S11 Threshold", S11,
     "Supplementary Table S11. Instrument-threshold sensitivity analysis. The primary analysis used exposure instruments selected at P<1e-5. The sensitivity analysis retained only genome-wide significant exposure instruments (P<5e-8) before harmonization and F>10 filtering."),
    ("S12 Threshold crosswalk", S12,
     "Supplementary Table S12. Phenotype-level cross-threshold comparison for phenotype records analyzable under both instrument thresholds (607 per outcome). For each immune phenotype and outcome, the GWAS ID, number of instruments, odds ratio with 95% confidence interval, P value, and FDR under the primary (P<1e-5) and genome-wide (P<5e-8) thresholds are shown side by side. These are the underlying data for Section 3.5 and Figure 3 and are sorted by P<5e-8."),
    ("S13 Chen re-analysis", S13,
     "Supplementary Table S13. Targeted re-analysis of the previously reported association between genetically predicted lymphocyte count and atrioventricular block (Chen Y et al, Front Immunol 2023;14:1041591), which used Blood Cell Consortium exposures and FinnGen release 2. Instruments here were taken from the same exposure source (GWAS Catalog GCST90002316; Chen MH et al, Cell 2020; 524,923 European-ancestry participants), clumped at r2<0.001 within 10 Mb against the 1000 Genomes European panel and filtered at F>10, and tested against FinnGen release 11. Release 11 contains the release 2 participants, so this is a re-analysis in an expanded release of the same cohort and not an independent validation; the two sets of estimates are not independent."),
    *POSTHOC_SHEETS,
    ("S17 Endpoint definitions", S17,
     "Supplementary Table S17. Definitions of the two FinnGen release 11 outcome endpoints. Case counts for the bundle-branch-block component endpoints are reported in Supplementary Table S23; a mutually exclusive decomposition of I9_CONDUCTIO into component endpoints remains unavailable."),
    ("S18 Power per phenotype", S18,
     "Supplementary Table S18. Instrument strength and statistical power for all 731 immunophenotypes, by outcome."),
    ("S19 Diagnostic coverage", S19,
     "Supplementary Table S19. Which sensitivity diagnostics are computable under each instrument threshold, with the number and proportion of phenotypes meeting the minimum instrument count for each, and the proportion showing significant Cochran's Q heterogeneity."),
    ("S20 Cochran Q per phenotype", S20,
     "Supplementary Table S20. Cochran's Q heterogeneity statistic for every immunophenotype for which it is defined (at least three instruments), under both instrument thresholds and for both outcomes, computed with conventional first-order weights and with modified second-order weights that account for uncertainty in the variant-exposure coefficients."),
    ("S21 Instrument QC sensitivity", S21,
     "Supplementary Table S21. Primary screen repeated after removing individual instruments whose reported effect exceeds one standard deviation per allele or whose expected minor-allele count is below 20, compared with the primary analysis using all instruments."),
    ("S22 AF specificity control", S22,
     "Supplementary Table S22. Specificity control for the lymphocyte-count re-analysis: the same instrument set applied to atrial fibrillation alongside the two conduction endpoints."),
    ("S23 Sub-endpoint screen", S23,
     "Supplementary Table S23. Phenotype-wide screen repeated on the bundle-branch-block component endpoints (I9_LBBB, I9_RBBB), which are narrower and more homogeneous than the broader conduction endpoint."),
]

wb = Workbook()
wb.remove(wb.active)
HFILL = PatternFill("solid", fgColor="1F4E79")
HFONT = Font(bold=True, color="FFFFFF")
CAPF  = Font(bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="F7FAFC")
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

for name, df, caption in SHEETS:
    ws = wb.create_sheet(title=name[:31])
    ncol = max(1, len(df.columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    cap = ws.cell(row=1, column=1, value=caption)
    cap.font = CAPF
    cap.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 42
    # header row
    for ci, col in enumerate(df.columns, start=1):
        c = ws.cell(row=2, column=ci, value=str(col))
        c.font = HFONT; c.fill = HFILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        for ci, v in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=(None if pd.isna(v) else v))
            cell.border = BORDER
            if ri % 2 == 1:
                cell.fill = ALT_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=(ci <= 2))
            if isinstance(v, float):
                cell.number_format = "0.000"
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions
    # column widths
    for ci, col in enumerate(df.columns, start=1):
        w = max(len(str(col)), *(len(str(x)) for x in df.iloc[:200, ci-1])) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 8), 48)

out = os.path.join(PKG, "Supplementary_Tables.xlsx")

# Carry over any sheets that this script does not build (for example S13, which is
# compiled from the output of 22_chen_reanalysis.R), so that re-running does not
# silently drop them. Captions and titles of carried-over sheets are declared here rather
# than edited into the workbook by hand, so that wording stays under version control.
CARRIED_RENAME = {}
CARRIED_CAPTION_FIX = {}
if os.path.exists(out):
    from openpyxl import load_workbook
    prev = load_workbook(out)
    for name in prev.sheetnames:
        new_name = CARRIED_RENAME.get(name, name)
        if new_name in wb.sheetnames:
            continue
        src = prev[name]
        dst = wb.create_sheet(title=new_name[:31])
        key = re.match(r"(S\d+)", new_name)
        fix = CARRIED_CAPTION_FIX.get(key.group(1)) if key else None
        for row in src.iter_rows():
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    if fix and cell.row == 1 and isinstance(value, str):
                        value = value.replace(*fix)
                    nc = dst.cell(row=cell.row, column=cell.column, value=value)
                    if cell.font and cell.font.bold:
                        nc.font = Font(bold=True)
        for letter, dim in src.column_dimensions.items():
            if dim.width:
                dst.column_dimensions[letter].width = dim.width

def sheet_number(ws):
    match = re.match(r"S(\d+)\b", ws.title)
    return int(match.group(1)) if match else 999

wb._sheets = sorted(wb._sheets, key=sheet_number)
wb.save(out)
print("written:", out)
for name, df, _ in SHEETS:
    print(f"  {name}: {len(df)} rows x {len(df.columns)} cols")
